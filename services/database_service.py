import pandas as pd  # Add pandas import at the top
import sqlite3
import os
import qrcode
from datetime import date, datetime, timedelta, timezone
from PIL import Image, ImageDraw, ImageFont
from bidi.algorithm import get_display  # Correctly display RTL text
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

QR_CODE_DIR = "qr_codes"

def store_session_token(token, expiry):
    print(f"DEBUG: Storing token {token} with expiry {expiry}")
    with sqlite3.connect("database.db") as conn:
        cursor = conn.cursor()
        cursor.execute('''
        INSERT INTO sessions (token, expiry) 
        VALUES (?, ?)
        ''', (token, expiry))
        conn.commit()
        return f"Token {token} stored with expiry {expiry}"


def validate_session_token(token):
    with sqlite3.connect("database.db") as conn:
        cursor = conn.cursor()
        cursor.execute('''
        SELECT expiry FROM sessions WHERE token = ?
        ''', (token,))
        result = cursor.fetchone()
        if result:
            expiry = datetime.fromisoformat(result[0])  # Token expiry in UTC
            current_time = datetime.now(timezone.utc)  # Current time in UTC
            print(f"DEBUG: Token expiry: {expiry}, Current time: {current_time}")
            if expiry > current_time:
                return True  # Token is valid
            else:
                print("DEBUG: Token has expired.")
    return False  # Token is invalid or expired


def remove_expired_tokens():
    with sqlite3.connect("database.db") as conn:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM sessions WHERE expiry < ?', (datetime.now(timezone.utc).isoformat(),))
        deleted_rows = cursor.rowcount  # Get the number of rows deleted
        conn.commit()
        print(f"Expired tokens cleaned up: {deleted_rows}")
        return deleted_rows


def get_books(order_by="desc"):
    with sqlite3.connect("database.db") as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Validate the order_by parameter
        order_clause = "DESC" if order_by.lower() == "desc" else "ASC"
        print(f"Order by: {order_clause}")

        # Query books with sorting and optional borrower details
        cursor.execute(f"""
            SELECT 
                books.*, 
                loans.borrowed_at, 
                members.parent_name AS borrowing_child,
                CASE WHEN loans.returned_at IS NULL THEN 'borrowed' ELSE 'available' END AS loan_status
            FROM books
            LEFT JOIN loans ON books.id = loans.book_id AND loans.returned_at IS NULL
            LEFT JOIN members ON loans.member_id = members.id
            ORDER BY books.created_at {order_clause}
        """)

        books = cursor.fetchall()
        print("Books fetched:", books)
        return [dict(book) for book in books]




def add_book(title, author, description, year_of_publication, cover_type, pages, recommended_age, book_condition, loan_status, delivering_parent):
    # Generate a unique QR code for the book
    with sqlite3.connect("database.db", timeout=5) as conn:  # Set timeout to 5 seconds
        cursor = conn.cursor()

        # Set to WAL mode for improved concurrency
        cursor.execute("PRAGMA journal_mode=WAL;")

        # Get the next available ID to use in QR code generation
        cursor.execute("SELECT COALESCE(MAX(id), 0) + 1 FROM books")
        new_id = cursor.fetchone()[0]
        qr_code = f"qr_for_book_{new_id}"

        # Generate and save the QR code image
        generate_qr_code_with_logo(qr_code, title)

        # Insert the book record into the database
        cursor.execute('''
            INSERT INTO books (qr_code, title, author, description, year_of_publication, cover_type, pages, recommended_age, book_condition, loan_status, delivering_parent)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (qr_code, title, author, description, year_of_publication, cover_type, pages, recommended_age, book_condition, loan_status, delivering_parent))

        # Commit the transaction to save changes
        conn.commit()

    return qr_code


def generate_qr_code_with_logo(qr_code, title):
    from bidi.algorithm import get_display
    import qrcode
    from PIL import Image, ImageDraw, ImageFont
    import os

    # Convert the Hebrew text to display correctly in RTL
    rtl_title = get_display(title)

    # Create a QR code with adjusted settings for longer text
    qr = qrcode.QRCode(
        version=None,  # Allow automatic version selection based on content
        error_correction=qrcode.constants.ERROR_CORRECT_Q,  # Changed to Q level for better balance
        box_size=12,  # Increased box size for better readability
        border=5,  # Slightly larger border
    )

    # Add data and optimize size
    qr.add_data(qr_code)
    qr.make(fit=True)

    # Get the QR code version that was selected
    current_version = qr.version
    print(f"QR Code Version: {current_version}")  # Debug info

    # Create QR code image with increased size
    qr_img = qr.make_image(fill="black", back_color="white").convert("RGB")

    # Get the base size of the QR code
    qr_base_size = qr_img.size[0]

    # Load and resize the logo
    logo_path = "./static/icm_logo.png"
    try:
        logo = Image.open(logo_path)
        # Reduce logo size to 20% of QR code (smaller than original 25%)
        logo_size = int(qr_base_size * 0.20)
        logo = logo.resize((logo_size, logo_size), Image.LANCZOS)

        # Create a white background for the logo to improve contrast
        logo_bg = Image.new('RGBA', (logo_size + 8, logo_size + 8), 'white')
        logo_pos = ((logo_bg.size[0] - logo_size) // 2, (logo_bg.size[1] - logo_size) // 2)
        logo_bg.paste(logo, logo_pos, mask=logo if logo.mode == 'RGBA' else None)

        # Calculate position for the logo
        pos = ((qr_base_size - logo_bg.size[0]) // 2, (qr_base_size - logo_bg.size[1]) // 2)

        # Create a mask for the logo area
        mask = Image.new('L', qr_img.size, 255)
        mask_draw = ImageDraw.Draw(mask)
        mask_draw.rectangle(
            [pos[0], pos[1], pos[0] + logo_bg.size[0], pos[1] + logo_bg.size[1]],
            fill=0
        )

        # Paste the logo with the white background
        qr_img.paste(logo_bg, pos, mask=logo_bg if logo_bg.mode == 'RGBA' else None)

    except Exception as e:
        print(f"Error adding logo to QR code: {e}")
        # Continue without logo if there's an error

    # Add space for title
    title_space = 50  # Increased space for title
    canvas = Image.new("RGB", (qr_img.size[0], qr_img.size[1] + title_space), "white")
    canvas.paste(qr_img, (0, 0))

    # Add title
    draw = ImageDraw.Draw(canvas)
    try:
        # Try to load Arial font, fallback to default if not available
        try:
            font = ImageFont.truetype("arial.ttf", 24)  # Increased font size
        except IOError:
            font = ImageFont.load_default()

        # Center the title text
        text_bbox = draw.textbbox((0, 0), rtl_title, font=font)
        text_width = text_bbox[2] - text_bbox[0]
        text_height = text_bbox[3] - text_bbox[1]
        text_position = (
            (canvas.size[0] - text_width) // 2,
            qr_img.size[1] + (title_space - text_height) // 2
        )

        # Draw text with a small white outline for better readability
        draw.text((text_position[0]-1, text_position[1]), rtl_title, fill="white", font=font)
        draw.text((text_position[0]+1, text_position[1]), rtl_title, fill="white", font=font)
        draw.text((text_position[0], text_position[1]-1), rtl_title, fill="white", font=font)
        draw.text((text_position[0], text_position[1]+1), rtl_title, fill="white", font=font)
        draw.text(text_position, rtl_title, fill="black", font=font)

    except Exception as e:
        print(f"Error adding title: {e}")

    # Save with high quality
    qr_code_dir = "qr_codes"
    if not os.path.exists(qr_code_dir):
        os.makedirs(qr_code_dir)

    output_path = os.path.join(qr_code_dir, f"{qr_code}.png")
    canvas.save(output_path, "PNG", quality=95)

    # Print debug info
    print(f"Generated QR code: {output_path}")
    print(f"QR code size: {qr_img.size}")
    print(f"Final image size: {canvas.size}")

    return output_path


def update_book_status(qr_code, status):
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("UPDATE books SET status = ? WHERE qr_code = ?", (status, qr_code))
    conn.commit()
    conn.close()

def get_book_loans(book_id):
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM loans WHERE book_id = ?", (book_id,))
    loans = cursor.fetchall()
    conn.close()
    return loans

def borrow_book(qr_code, member_id, borrowed_date, book_state):
    with sqlite3.connect("database.db") as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Lookup book_id from qr_code
        cursor.execute("SELECT id FROM books WHERE qr_code = ?", (qr_code,))
        book = cursor.fetchone()

        if not book:
            return False  # Book with given qr_code not found

        book_id = book["id"]

        # Insert a new loan record with book_id, member_id, borrowed_date, and book_state
        cursor.execute("""
            INSERT INTO loans (book_id, member_id, borrowed_at, book_state)
            VALUES (?, ?, ?, ?)
        """, (book_id, member_id, borrowed_date, book_state))

        # Update the loan_status in the books table to 'borrowed'
        cursor.execute("""
            UPDATE books SET loan_status = 'borrowed' WHERE id = ?
        """, (book_id,))

        conn.commit()

    return True


def update_book(book_id, **kwargs):
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    try:
        # Filter allowed fields and prepare update parameters
        allowed_fields = {
            'title', 'author', 'description', 'year_of_publication',
            'cover_type', 'pages', 'recommended_age', 'book_condition',
            'delivering_parent'
        }

        update_fields = []
        values = []

        for field, value in kwargs.items():
            if field in allowed_fields:
                update_fields.append(f"{field} = ?")
                values.append(value)

        if not update_fields:
            return None  # No valid fields to update

        # Add book_id as the last parameter
        values.append(book_id)

        # Build the update query
        query = f'''
            UPDATE books 
            SET {', '.join(update_fields)}
            WHERE id = ?
        '''

        cursor.execute(query, values)
        conn.commit()

        # Return the updated book
        cursor.execute('''
            SELECT id, title, author, description, year_of_publication, cover_type, 
                   pages, recommended_age, book_condition, loan_status, delivering_parent, qr_code
            FROM books 
            WHERE id = ?
        ''', (book_id,))

        book = cursor.fetchone()
        if book:
            return {
                "id": book[0],
                "title": book[1],
                "author": book[2],
                "description": book[3],
                "year_of_publication": book[4],
                "cover_type": book[5],
                "pages": book[6],
                "recommended_age": book[7],
                "book_condition": book[8],
                "loan_status": book[9],
                "delivering_parent": book[10],
                "qr_code": book[11]
            }
        return None

    except sqlite3.Error as e:
        print(f"Database error: {str(e)}")
        conn.rollback()
        return None
    finally:
        conn.close()


def return_book(qr_code):
    with sqlite3.connect("database.db") as conn:
        cursor = conn.cursor()

        # Find the book ID from the QR code
        cursor.execute("SELECT id FROM books WHERE qr_code = ?", (qr_code,))
        book = cursor.fetchone()

        if not book:
            return {"success": False, "message": "Book not found"}

        book_id = book[0]

        # Update the returned_at timestamp for the latest loan for this book
        cursor.execute("""
            UPDATE loans 
            SET returned_at = CURRENT_TIMESTAMP 
            WHERE book_id = ? AND returned_at IS NULL
        """, (book_id,))

        # Check if there are any other active loans for this book
        cursor.execute("""
            SELECT COUNT(*) FROM loans WHERE book_id = ? AND returned_at IS NULL
        """, (book_id,))
        active_loans_count = cursor.fetchone()[0]

        # If there are no active loans, update the loan_status in the books table to 'available'
        if active_loans_count == 0:
            cursor.execute("""
                UPDATE books SET loan_status = 'available' WHERE id = ?
            """, (book_id,))

        conn.commit()

    return {"success": True, "message": "Book returned successfully"}


def get_members():
    with sqlite3.connect("database.db") as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM members ORDER BY created_at DESC")
        rows = cursor.fetchall()
        return [dict(row) for row in rows]

def add_member(parent_name, kid_name, email):
    with sqlite3.connect("database.db") as conn:
        cursor = conn.cursor()
        cursor.execute('''
        INSERT INTO members (parent_name, kid_name, email)
        VALUES (?, ?, ?)
        ''', (parent_name, kid_name, email))
        conn.commit()

def update_member(member_id, parent_name, kid_name, email):
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE members 
        SET parent_name = ?, kid_name = ?, email = ?
        WHERE id = ?
    ''', (parent_name, kid_name, email, member_id))
    conn.commit()
    conn.close()

def delete_member(member_id):
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    # Check if there are any open loans for this member
    cursor.execute(
        "SELECT COUNT(*) FROM loans WHERE member_id = ? AND returned_at IS NULL",
        (member_id,)
    )
    open_loans_count = cursor.fetchone()[0]

    if open_loans_count > 0:
        conn.close()
        # Instead of deleting, we raise an exception.
        raise Exception("Cannot delete member with open loans.")

    # Proceed to delete the member if no open loans
    cursor.execute('DELETE FROM members WHERE id = ?', (member_id,))
    conn.commit()
    conn.close()

def get_book_by_qr_code(qr_code):
    with sqlite3.connect("database.db") as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM books WHERE qr_code = ?", (qr_code,))
        book = cursor.fetchone()
    return dict(book) if book else None


def get_books_by_status(param):
    with sqlite3.connect("database.db") as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM books WHERE loan_status = ?", (param,))
        rows = cursor.fetchall()
        books = [dict(row) for row in rows]
    return books

def get_borrowing_history(qr_code=None):
    query = '''
    SELECT 
        books.title AS book_title,
        loans.book_id AS book_id,
        members.kid_name AS borrowed_name,
        loans.borrowed_at AS loan_start,
        loans.returned_at AS return_date,
        loans.book_state AS state
    FROM loans
    JOIN books ON books.id = loans.book_id
    JOIN members ON members.id = loans.member_id
    '''
    params = []
    if qr_code:
        query += " WHERE books.qr_code = ?"
        params.append(qr_code)

    query += " ORDER BY loans.borrowed_at DESC"

    with sqlite3.connect("database.db") as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute(query, params)
        results = cursor.fetchall()

    return [dict(row) for row in results]

def get_open_loans(qr_code=None):
    query = """
        SELECT 
            l.id AS loan_id,
            l.book_id,
            b.title AS book_title,
            m.parent_name AS borrower_name,
            l.borrowed_at AS loan_start_date,
            l.returned_at AS return_date
        FROM loans l
        JOIN books b ON l.book_id = b.id
        JOIN members m ON l.member_id = m.id
        WHERE l.returned_at IS NULL
    """
    params = []
    if qr_code:
        query += " AND b.qr_code = ?"
        params.append(qr_code)
    with sqlite3.connect("database.db") as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute(query, params)
        return [dict(row) for row in cursor.fetchall()]


def get_loan_history(qr_code, show_all):
    with sqlite3.connect("database.db") as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        query = """
            SELECT l.id, l.book_id, l.borrowed_at, l.returned_at, l.book_state, b.title AS book_title, m.parent_name AS borrower_name, m.kid_name AS borrower_child
            FROM loans l
            JOIN books b ON l.book_id = b.id
            JOIN members m ON l.member_id = m.id
            WHERE b.qr_code = ?
        """
        if not show_all:
            query += " AND l.returned_at IS NULL"  # Filter for open loans only
        query += " ORDER BY l.borrowed_at DESC"

        cursor.execute(query, (qr_code,))
        loans = cursor.fetchall()

    return [dict(loan) for loan in loans]


def get_all_open_loans():
    print("getting only open loans")
    with sqlite3.connect("database.db") as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        query = """
            SELECT l.id, l.book_id, l.borrowed_at, l.returned_at, l.book_state, b.title AS book_title, m.parent_name AS borrower_name, m.kid_name AS borrower_child
            FROM loans l
            JOIN books b ON l.book_id = b.id
            JOIN members m ON l.member_id = m.id
            WHERE l.returned_at IS NULL  -- Only open loans
            ORDER BY l.borrowed_at DESC
        """
        cursor.execute(query)
        loans = cursor.fetchall()

    return [dict(loan) for loan in loans]

def get_all_loans():
    print("getting all loans")
    with sqlite3.connect("database.db") as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        query = """
            SELECT l.id, l.book_id, l.borrowed_at, l.returned_at, l.book_state, b.title AS book_title, m.parent_name AS borrower_name, m.kid_name AS borrower_child
            FROM loans l
            JOIN books b ON l.book_id = b.id
            JOIN members m ON l.member_id = m.id
            ORDER BY l.borrowed_at DESC
        """
        cursor.execute(query)
        loans = cursor.fetchall()

    return [dict(loan) for loan in loans]


# Function to extract book data for reporting
def get_books_report(order_by="desc", sort_column="title", include_history=True):
    with sqlite3.connect("database.db") as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Validate the order_by parameter
        order_clause = "DESC" if order_by.lower() == "desc" else "ASC"
        valid_sort_columns = ["created_at", "borrowed_at", "title"]
        sort_column = sort_column if sort_column in valid_sort_columns else "title"

        # Fetch loan data based on include_history parameter
        if include_history:
            # Fetch all loans (both open and closed)
            query = f"""
                SELECT 
                    books.title,
                    books.author,
                    loans.borrowed_at,
                    loans.returned_at,
                    members.parent_name AS borrowed_by,
                    members.email AS borrower_email
                FROM books
                LEFT JOIN loans ON books.id = loans.book_id
                LEFT JOIN members ON loans.member_id = members.id
                ORDER BY books.{sort_column} {order_clause}
            """
        else:
            # Fetch only open loans
            query = f"""
                SELECT 
                    books.title,
                    books.author,
                    loans.borrowed_at,
                    members.parent_name AS borrowed_by,
                    members.email AS borrower_email
                FROM books
                LEFT JOIN loans ON books.id = loans.book_id AND loans.returned_at IS NULL
                LEFT JOIN members ON loans.member_id = members.id
                WHERE loans.returned_at IS NULL
                ORDER BY books.{sort_column} {order_clause}
            """

        cursor.execute(query)
        rows = cursor.fetchall()
        return [dict(row) for row in rows]

# Function to generate an Excel report
def generate_books_report(order_by="desc", sort_column="title", include_history=True, language="he"):
    # Extract the data using the `get_books_report` function
    books_data = get_books_report(order_by, sort_column, include_history)

    # Create a DataFrame from the extracted data
    df = pd.DataFrame(books_data)

    # **Dynamic Header Translation**
    header_translations = {
        'en': {
            'title': 'Title',
            'author': 'Author',
            'borrowed_at': 'Borrowed At',
            'returned_at': 'Returned At',
            'borrowed_by': 'Borrowed By',
            'borrower_email': 'Borrower Email'
        },
        'he': {
            'title': 'שם הספר',
            'author': 'שם המחבר',
            'borrowed_at': 'תאריך השאלה',
            'returned_at': 'תאריך החזרה',
            'borrowed_by': 'הושאל על ידי',
            'borrower_email': 'אימייל השואל'
        }
    }

    # Get headers based on the language
    headers = header_translations.get(language, header_translations['en'])
    translated_columns = [headers.get(col, col) for col in df.columns]
    df.columns = translated_columns  # Rename DataFrame columns to translated headers

    if language == 'he':
        df.columns = [get_display(col) for col in df.columns]  # Apply RTL for Hebrew

    # Generate the Excel file
    report_type = "Active Loans" if not include_history else "Historical Loans"
    report_filename = f"books_report_{report_type}_{date.today()}.xlsx"

    # Create a workbook and add the data
    wb = Workbook()
    ws = wb.active
    ws.title = "ספריית הקהילה הישראלית במדריד - " if language == 'he' else "ICM Library - "

    # Set header
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4ECAC7", end_color="4ECAC7", fill_type="solid")
    report_title = "דוח השאלות" if language == 'he' else "Loans Report"
    ws.append([report_title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    header_cell = ws.cell(row=1, column=1)
    header_cell.font = Font(bold=True, size=14)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows with borders
    border = Border(left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000"))

    # Add headers
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 2):
        if r_idx == 2:
            for c_idx, cell_value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=cell_value)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
        else:
            for c_idx, cell_value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=cell_value)
                cell.border = border
                if include_history and df.at[r_idx - 3, 'returned_at'] is None:
                    cell.fill = PatternFill(start_color="FEC43C", end_color="FEC43C", fill_type="solid")

    # Freeze the header row
    ws.freeze_panes = "A3"

    # Add filters to columns (start from the second row)
    ws.auto_filter.ref = f"A2:{ws.cell(row=2, column=len(df.columns)).coordinate}"

    # Save the workbook
    wb.save(report_filename)

    return report_filename

# Function to generate an inventory report
def generate_inventory_report(order_by="desc", sort_column="title", include_borrowed=True, language="he"):
    with sqlite3.connect("database.db") as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        print("include_borrowed:", include_borrowed)

        # Validate the order_by parameter
        order_clause = "DESC" if order_by.lower() == "desc" else "ASC"
        valid_sort_columns = ["created_at", "title"]
        sort_column = sort_column if sort_column in valid_sort_columns else "title"

        # Fetch inventory data based on include_borrowed parameter
        if include_borrowed:
            query = f"""
                SELECT 
                    books.id,
                    books.title,
                    books.author,
                    books.description,
                    books.year_of_publication,
                    books.pages,
                    books.cover_type,
                    books.book_condition,
                    books.loan_status
                FROM books
                ORDER BY books.{sort_column} {order_clause}
            """
        else:
            query = f"""
                SELECT 
                    books.id,
                    books.title,
                    books.author,
                    books.description,
                    books.year_of_publication,
                    books.pages,
                    books.cover_type,
                    books.book_condition,
                    books.loan_status
                FROM books
                WHERE books.loan_status = 'available'
                ORDER BY books.{sort_column} {order_clause}
            """

        cursor.execute(query)
        rows = cursor.fetchall()
        books_data = [dict(row) for row in rows]

    # Create a DataFrame from the extracted data
    df = pd.DataFrame(books_data)

    # **Dynamic Header Translation**
    header_translations = {
        'en': {
            'id': 'ID',
            'title': 'Title',
            'author': 'Author',
            'description': 'Description',
            'year_of_publication': 'Year of Publication',
            'pages': 'Pages',
            'cover_type': 'Cover Type',
            'book_condition': 'Book Condition',
            'loan_status': 'Loan Status'
        },
        'he': {
            'id': 'מזהה',
            'title': 'שם הספר',
            'author': 'שם המחבר',
            'description': 'תיאור',
            'year_of_publication': 'שנת פרסום',
            'pages': 'עמודים',
            'cover_type': 'סוג הכריכה',
            'book_condition': 'מצב הספר',
            'loan_status': 'סטטוס השאלה'
        }
    }

    # Get headers based on the language
    headers = header_translations.get(language, header_translations['en'])
    translated_columns = [headers.get(col, col) for col in df.columns]
    df.columns = translated_columns  # Rename DataFrame columns to translated headers

    if language == 'he':
        df.columns = [get_display(col) for col in df.columns]  # Apply RTL for Hebrew

    # Generate the Excel file
    report_filename = f"inventory_report_{date.today()}.xlsx"

    # Create a workbook and add the data
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Add report title
    report_title = "דוח מלאי" if language == 'he' else "Inventory Report"
    ws.append([report_title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    header_cell = ws.cell(row=1, column=1)
    header_cell.font = Font(bold=True, size=14)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows with borders
    border = Border(left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000"))

    # Set header
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4ECAC7", end_color="4ECAC7", fill_type="solid")
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 2):
        if r_idx == 2:  # Headers row
            for c_idx, cell_value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=cell_value)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
        else:  # Data rows
            for c_idx, cell_value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=cell_value)
                cell.border = border

    # Freeze the header row
    ws.freeze_panes = "A3"

    # Add filters to columns (start from the second row)
    ws.auto_filter.ref = f"A2:{ws.cell(row=2, column=len(df.columns)).coordinate}"

    # Save the workbook
    wb.save(report_filename)

    return report_filename


def find_email_by_borrower_name(borrower_name):
    """
    Look up the email of a member by the borrower name.
    The borrower name corresponds to the `kid_name` in the members table.
    """
    try:
        with sqlite3.connect("database.db") as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT email FROM members WHERE parent_name = ?", (borrower_name,))
            result = cursor.fetchone()
            if result:
                email = result["email"]  # Extract the email from the result
                return email
            else:
                print(f"No email found for borrower name: {borrower_name}")
                return None
    except Exception as e:
        print(f"Error finding email for borrower name '{borrower_name}': {e}")
        return None


def check_recent_reminder(loan_id, days=14):
    """
    Check if a reminder has been sent for this loan in the past 'days' days.
    """
    with sqlite3.connect("database.db") as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cutoff_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
        cursor.execute('''
        SELECT 1 FROM reminders 
        WHERE loan_id = ? AND sent_at >= ?
        ''', (loan_id, cutoff_date))
        result = cursor.fetchone()
        return result is not None


def record_reminder(loan_id):
    """
    Record that a reminder has been sent for a specific loan.
    """
    with sqlite3.connect("database.db") as conn:
        cursor = conn.cursor()
        sent_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        try:
            cursor.execute('''
            INSERT INTO reminders (loan_id, sent_at) 
            VALUES (?, ?)
            ''', (loan_id, sent_at))
            conn.commit()
        except Exception as e:
            print(f"Failed to insert reminder record for loan_id {loan_id}: {e}")


def fetch_last_reminder_date(loan_id):
    """
    Get the most recent reminder date for a specific loan_id.
    """
    with sqlite3.connect("database.db") as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('''
        SELECT sent_at 
        FROM reminders 
        WHERE loan_id = ? 
        ORDER BY sent_at DESC 
        LIMIT 1
        ''', (loan_id,))
        result = cursor.fetchone()
        return result["sent_at"] if result else None